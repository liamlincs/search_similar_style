import argparse
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "src"
import sys
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from color_card_store import ColorCardStore


DEFAULT_FILES = [
    "彩龙丝光棉（2024）.xlsx",
    "东莞国彩丝光棉.xlsx",
    "恩盛纺织（新色卡）丝光棉.xlsx",
]


def slugify(value: str) -> str:
    mapping = {
        "彩龙丝光棉（2024）": "cailong_2024",
        "东莞国彩丝光棉": "dongguan_guocai",
        "恩盛纺织（新色卡）丝光棉": "ensheng_new",
    }
    if value in mapping:
        return mapping[value]
    raw = re.sub(r"\.[^.]+$", "", value).strip().lower()
    raw = re.sub(r"[^a-z0-9]+", "_", raw)
    return raw.strip("_") or "library"


def read_color_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb["result"] if "result" in wb.sheetnames else wb.worksheets[0]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    header_idx = {name: idx + 1 for idx, name in enumerate(headers)}
    required = ["名称", "L", "a", "b"]
    missing = [name for name in required if name not in header_idx]
    if missing:
        raise ValueError(f"{path.name} missing columns: {missing}")

    wave_cols = []
    for idx, name in enumerate(headers, 1):
        if re.fullmatch(r"\d+nm", name):
            wave_cols.append((int(name[:-2]), idx))

    rows = []
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, header_idx["名称"]).value or "").strip()
        if not name:
            continue
        try:
            l_val = float(ws.cell(r, header_idx["L"]).value)
            a_val = float(ws.cell(r, header_idx["a"]).value)
            b_val = float(ws.cell(r, header_idx["b"]).value)
        except (TypeError, ValueError):
            continue
        spectral = []
        for wavelength, c in wave_cols:
            value = ws.cell(r, c).value
            if value is None:
                continue
            try:
                spectral.append({"wavelength": wavelength, "value": float(value)})
            except (TypeError, ValueError):
                pass
        rows.append(
            {
                "name": name,
                "note": ws.cell(r, header_idx.get("备注", 0)).value if header_idx.get("备注") else "",
                "illuminant": ws.cell(r, header_idx.get("光源", 0)).value if header_idx.get("光源") else "",
                "angle": ws.cell(r, header_idx.get("角度", 0)).value if header_idx.get("角度") else None,
                "l": l_val,
                "a": a_val,
                "b": b_val,
                "spectral": spectral,
            }
        )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Import color-card xlsx files into SQLite.")
    parser.add_argument("--db", default=str(ROOT / "data" / "color_cards.db"))
    parser.add_argument("--source-dir", default=str(ROOT / "tools" / "ColorMeter_miniprogram_bluetooth_example"))
    parser.add_argument("files", nargs="*", default=DEFAULT_FILES)
    args = parser.parse_args()

    store = ColorCardStore(Path(args.db))
    source_dir = Path(args.source_dir)
    total = 0
    for order, filename in enumerate(args.files):
        path = Path(filename)
        if not path.is_absolute():
            path = source_dir / filename
        library_name = path.stem
        library_id = slugify(library_name)
        rows = read_color_rows(path)
        count = store.replace_library(library_id, library_name, path.name, rows, sort_order=order)
        total += count
        print(f"{library_id}\t{library_name}\t{count}")
    print(f"total\t{total}")


if __name__ == "__main__":
    main()

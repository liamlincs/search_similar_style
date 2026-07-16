import re
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

import openpyxl

from color_card_store import clean_library_id


KNOWN_LIBRARY_IDS = {
    "彩龙丝光棉（2024）": "cailong_2024",
    "东莞国彩丝光棉": "dongguan_guocai",
    "恩盛纺织（新色卡）丝光棉": "ensheng_new",
}


def slugify_library_id(value: str) -> str:
    raw_value = str(value or "").strip()
    if raw_value in KNOWN_LIBRARY_IDS:
        return KNOWN_LIBRARY_IDS[raw_value]
    ascii_raw = re.sub(r"\.[^.]+$", "", raw_value).strip().lower()
    ascii_raw = re.sub(r"[^a-z0-9]+", "_", ascii_raw).strip("_")
    return ascii_raw or clean_library_id(raw_value) or "library"


def _cell_value(ws: Any, row: int, col: int | None) -> Any:
    if not col:
        return ""
    return ws.cell(row, col).value


def read_color_rows(source: Path | bytes | BinaryIO) -> list[dict[str, Any]]:
    if isinstance(source, Path):
        workbook_source: Path | BytesIO = source
        source_name = source.name
    elif isinstance(source, bytes):
        workbook_source = BytesIO(source)
        source_name = "uploaded workbook"
    else:
        workbook_source = source
        source_name = "uploaded workbook"

    wb = openpyxl.load_workbook(workbook_source, read_only=False, data_only=True)
    ws = wb["result"] if "result" in wb.sheetnames else wb.worksheets[0]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    header_idx = {name: idx + 1 for idx, name in enumerate(headers) if name}
    required = ["名称", "L", "a", "b"]
    missing = [name for name in required if name not in header_idx]
    if missing:
        raise ValueError(f"{source_name} 缺少列：{', '.join(missing)}")

    wave_cols: list[tuple[int, int]] = []
    for idx, name in enumerate(headers, 1):
        if re.fullmatch(r"\d+nm", name):
            wave_cols.append((int(name[:-2]), idx))

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, header_idx["名称"]).value or "").strip()
        if not name:
            continue
        try:
            l_val = float(ws.cell(r, header_idx["L"]).value)
            a_val = float(ws.cell(r, header_idx["a"]).value)
            b_val = float(ws.cell(r, header_idx["b"]).value)
        except (TypeError, ValueError):
            continue
        spectral = []
        for wavelength, c in wave_cols:
            value = ws.cell(r, c).value
            if value is None:
                continue
            try:
                spectral.append({"wavelength": wavelength, "value": float(value)})
            except (TypeError, ValueError):
                pass
        rows.append(
            {
                "name": name,
                "note": _cell_value(ws, r, header_idx.get("备注")),
                "illuminant": _cell_value(ws, r, header_idx.get("光源")),
                "angle": _cell_value(ws, r, header_idx.get("角度")),
                "l": l_val,
                "a": a_val,
                "b": b_val,
                "spectral": spectral,
            }
        )
    return rows
